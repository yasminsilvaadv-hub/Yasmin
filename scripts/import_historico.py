#!/usr/bin/env python3
"""
Script de importação do histórico real da Inext Holding.
Lê o Excel e sobe no Supabase:
  - Ativos (ON, PNA, PN)
  - Pessoas (stakeholders)
  - Eventos (AGE, AGC, ...)
  - Operações de ativos (operacoes_ativos)
  - Plano de equity + contratos
"""
import warnings
warnings.filterwarnings('ignore')

import sys
import openpyxl
from datetime import datetime, date
from supabase import create_client

# ─── Config ──────────────────────────────────────────────────────────────────
URL  = 'https://zewjorcoxoepcywnmypw.supabase.co'
KEY  = 'sb_secret_7p85nn-gLCnB9qT-_Ibeng_No2jiFLH'
XLSX = '/Users/yas/Downloads/all_operations_21-04-2026_Inext Holding de Participações em Tecnologia SA.xlsx'
ORG_SLUG = 'inext-holding-de-participacoes-em-tecnologia'

sb = create_client(URL, KEY)

# ─── Helpers ─────────────────────────────────────────────────────────────────
def fmt_date(v):
    if v is None: return None
    if isinstance(v, (datetime, date)): return v.strftime('%Y-%m-%dT00:00:00')
    return str(v)

def norm(s):
    if s is None: return None
    return str(s).strip()

# ─── 1. Org ───────────────────────────────────────────────────────────────────
org = sb.table('organizacoes').select('id,nome,slug').eq('slug', ORG_SLUG).single().execute().data
ORG_ID = org['id']
print(f"✓ Org: {org['nome']} [{ORG_ID}]")

# ─── 2. Ativos ────────────────────────────────────────────────────────────────
existing_ativos = {a['codigo']: a for a in
    sb.table('ativos').select('*').eq('organizacao_id', ORG_ID).execute().data}

def upsert_ativo(codigo, especie, nome_classe=None):
    if codigo in existing_ativos:
        # Corrige se necessário
        a = existing_ativos[codigo]
        if a.get('nome_classe') != nome_classe or a.get('especie') != especie:
            sb.table('ativos').update({'especie': especie, 'nome_classe': nome_classe}).eq('id', a['id']).execute()
            print(f"  ↻ Ativo atualizado: {codigo}")
        return a['id']
    r = sb.table('ativos').insert({
        'organizacao_id': ORG_ID,
        'tipo': 'acao',
        'codigo': codigo,
        'especie': especie,
        'nome_classe': nome_classe,
    }).execute()
    print(f"  + Ativo criado: {codigo}")
    return r.data[0]['id']

ativo_ids = {
    'ON':  upsert_ativo('ON',  'ordinaria',   None),
    'PNA': upsert_ativo('PNA', 'preferencial', 'Classe A'),
    'PN':  upsert_ativo('PN',  'preferencial', 'Classe B'),
}
print(f"✓ Ativos: {ativo_ids}")

# ─── 3. Pessoas ───────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws_acoes = wb['Ações']

all_people = set()
for row in ws_acoes.iter_rows(min_row=6, values_only=True):
    if row[0] is None: continue
    origem  = norm(row[5])
    destino = norm(row[6])
    if origem:  all_people.add(origem)
    if destino: all_people.add(destino)

ws_equity = wb['Equity plans']
for row in ws_equity.iter_rows(min_row=6, values_only=True):
    if row[0] is None: continue
    benef = norm(row[4])
    if benef: all_people.add(benef)

existing_pessoas = {p['nome_completo']: p['id'] for p in
    sb.table('pessoas').select('id,nome_completo').eq('organizacao_id', ORG_ID).execute().data}

pessoas_ids = dict(existing_pessoas)
novos = 0
for nome in sorted(all_people):
    if nome in pessoas_ids: continue
    # Heurística: PJ se contém Ltda, S.A., Fundo, FIP, Holding, etc.
    pj_keywords = ['ltda','s/a','s.a.','fundo','fip','holding','participações','participacoes','investimentos','inovação','inovacao','tecnologia']
    is_pj = any(k in nome.lower() for k in pj_keywords)
    tipo = 'pessoa_juridica' if is_pj else 'pessoa_fisica'
    r = sb.table('pessoas').insert({
        'organizacao_id': ORG_ID,
        'nome_completo': nome,
        'tipo': tipo,
    }).execute()
    pessoas_ids[nome] = r.data[0]['id']
    novos += 1
    print(f"  + Pessoa: {nome} [{tipo}]")

print(f"✓ Pessoas: {len(pessoas_ids)} total ({novos} novas)")

# ─── 4. Eventos ───────────────────────────────────────────────────────────────
# Agrupa (nome_evento, data_evento, deliberacao) únicos
eventos_set = {}
for row in ws_acoes.iter_rows(min_row=6, values_only=True):
    if row[0] is None: continue
    nome_ev  = norm(row[10])
    data_ev  = fmt_date(row[11])
    deliber  = norm(row[12])
    if nome_ev and nome_ev != '-' and data_ev:
        key = (nome_ev, data_ev)
        if key not in eventos_set:
            eventos_set[key] = deliber

existing_eventos = sb.table('eventos').select('id,nome,data_hora').eq('organizacao_id', ORG_ID).execute().data
evento_lookup = {(e['nome'], e['data_hora'][:10] if e['data_hora'] else ''): e['id'] for e in existing_eventos}

evento_ids = dict(evento_lookup)
for (nome_ev, data_ev), deliber in eventos_set.items():
    key = (nome_ev, data_ev[:10] if data_ev else '')
    if key in evento_ids: continue
    tipo_ev = 'age' if nome_ev.upper().startswith('AGE') else \
              'ago' if nome_ev.upper().startswith('AGO') else \
              'rd'  if nome_ev.upper() in ('AGC','AGCC') else 'rd'
    r = sb.table('eventos').insert({
        'organizacao_id': ORG_ID,
        'nome': nome_ev,
        'tipo': tipo_ev,
        'data_hora': data_ev,
        'status': 'concluido',
        'ordem_do_dia': deliber,
    }).execute()
    ev_id = r.data[0]['id']
    evento_ids[key] = ev_id
    print(f"  + Evento: {nome_ev} [{data_ev[:10]}]")

print(f"✓ Eventos: {len(evento_ids)}")

# ─── 5. Mapeamento de tipo de operação ───────────────────────────────────────
# Tipos no Excel → tipos no banco (+ metadata para rastreabilidade)
# A cap-table funciona corretamente com este mapeamento:
#   Subscrição      → emissao         (credita destino_id)
#   Transferência   → transferencia   (debita origem, credita destino)
#   Conversão IN    → emissao         (credita destino_id no novo ativo)
#   Conversão OUT   → cancelamento    (debita origem_id no ativo antigo)
#   Constituição    → onus_constituicao
#   Extinção        → onus_extincao

def map_tipo(operacao_excel, origem, destino):
    op = (operacao_excel or '').lower()
    if 'subscrição' in op or 'subscricao' in op:
        return 'emissao', 'subscricao'
    if 'transferência' in op or 'transferencia' in op:
        return 'transferencia', 'transferencia'
    if 'conversão' in op or 'conversao' in op:
        if destino and not origem:
            return 'emissao', 'conversao'
        else:
            return 'cancelamento', 'conversao'
    if 'constituição' in op or 'constituicao' in op or 'ônus' in op:
        if 'extinção' in op or 'extincao' in op:
            return 'onus_extincao', 'onus_extincao'
        return 'onus_constituicao', 'onus_constituicao'
    if 'extinção' in op or 'extincao' in op:
        return 'onus_extincao', 'onus_extincao'
    # fallback
    return 'emissao', op

# ─── 6. Operações de ativos ───────────────────────────────────────────────────
ABREV_MAP = {'ON': 'ON', 'PNA': 'PNA', 'PN': 'PN'}

erros = []
inseridas = 0

for i, row in enumerate(ws_acoes.iter_rows(min_row=6, values_only=True), start=6):
    if row[0] is None: continue

    especie_raw = norm(row[0])  # Ordinária / Preferencial
    abrev       = norm(row[2])  # ON / PNA / PN
    data_op     = fmt_date(row[3])
    operacao    = norm(row[4])
    origem_nome = norm(row[5])
    destino_nome= norm(row[6])
    quantidade  = row[7]
    valor       = row[8]        # Valor da operação
    nome_ev     = norm(row[10])
    data_ev     = fmt_date(row[11])
    deliber     = norm(row[12])
    anotacoes   = norm(row[14])

    if not abrev or not data_op or not quantidade:
        continue

    ativo_id = ativo_ids.get(abrev)
    if not ativo_id:
        erros.append(f"Linha {i}: ativo desconhecido '{abrev}'")
        continue

    tipo_banco, tipo_original = map_tipo(operacao, origem_nome, destino_nome)

    origem_id  = pessoas_ids.get(origem_nome)  if origem_nome  else None
    destino_id = pessoas_ids.get(destino_nome) if destino_nome else None

    # Preço unitário (valor/quantidade se valor > 0)
    preco = None
    if valor and quantidade and float(quantidade) > 0 and float(valor) > 0:
        preco = round(float(valor) / float(quantidade), 6)

    # Evento
    ev_key = (nome_ev, data_ev[:10] if data_ev else '') if (nome_ev and nome_ev != '-') else None
    ev_id  = evento_ids.get(ev_key) if ev_key else None

    metadata = {
        'tipo_original': tipo_original,
        'evento_id': str(ev_id) if ev_id else None,
        'deliberacao': deliber,
    }
    if anotacoes: metadata['anotacoes'] = anotacoes[:500]

    try:
        sb.table('operacoes_ativos').insert({
            'organizacao_id': ORG_ID,
            'ativo_id':       ativo_id,
            'tipo_operacao':  tipo_banco,
            'quantidade':     float(quantidade),
            'data_operacao':  data_op,
            'origem_id':      origem_id,
            'destino_id':     destino_id,
            'preco_unitario': preco,
            'motivo':         anotacoes[:200] if anotacoes else None,
            'metadata':       metadata,
        }).execute()
        inseridas += 1
    except Exception as e:
        erros.append(f"Linha {i} ({operacao} {abrev} {data_op}): {e}")

print(f"\n✓ Operações inseridas: {inseridas}")
if erros:
    print(f"✗ Erros ({len(erros)}):")
    for e in erros: print(f"   {e}")

# ─── 7. Equity Plans ─────────────────────────────────────────────────────────
print("\n─── Equity Plans ────────────────────────────────────────────────────────")

# Busca ativo ON (para stock options)
ativo_on_id = ativo_ids['ON']

# Cria ou encontra o plano
plano_nome = 'Plano de Opção de Compra de Ações - 01/2024'
existing_planos = sb.table('planos_equity').select('id,nome').eq('organizacao_id', ORG_ID).execute().data
plano_id = next((p['id'] for p in existing_planos if p['nome'] == plano_nome), None)
if not plano_id:
    r = sb.table('planos_equity').insert({
        'organizacao_id': ORG_ID,
        'nome': plano_nome,
        'tipo': 'stock_options',
        'ativo_id': ativo_on_id,
        'data_inicio': '2024-03-06',
    }).execute()
    plano_id = r.data[0]['id']
    print(f"  + Plano criado: {plano_nome}")
else:
    print(f"  ↻ Plano já existe: {plano_nome}")

# Programas únicos
programas_set = set()
for row in ws_equity.iter_rows(min_row=6, values_only=True):
    if row[0] is None: continue
    prog = norm(row[2])
    if prog: programas_set.add(prog)

existing_progs = sb.table('programas_equity').select('id,nome').eq('plano_id', plano_id).execute().data
prog_ids = {p['nome']: p['id'] for p in existing_progs}

for prog_nome in programas_set:
    if prog_nome in prog_ids: continue
    r = sb.table('programas_equity').insert({
        'plano_id': plano_id,
        'nome': prog_nome,
    }).execute()
    prog_ids[prog_nome] = r.data[0]['id']
    print(f"  + Programa: {prog_nome}")

# Contratos — agrupa por (beneficiario, contrato_cod, programa)
# A planilha mostra linhas de operações (Outorga / Finalização) por contrato
contratos_raw = {}
for row in ws_equity.iter_rows(min_row=6, values_only=True):
    if row[0] is None: continue
    prog_nome  = norm(row[2])
    contrato_c = norm(row[3])   # SO - 47, etc.
    benef_nome = norm(row[4]).strip().rstrip(',').strip() if row[4] else None
    status_raw = norm(row[5])   # Em curso / Finalizado
    data_op    = fmt_date(row[6])
    operacao   = norm(row[7])   # Outorga / Finalização
    qtd        = row[8]

    if not benef_nome or not contrato_c: continue

    key = (contrato_c, benef_nome, prog_nome)
    if key not in contratos_raw:
        contratos_raw[key] = {
            'prog': prog_nome,
            'benef': benef_nome,
            'codigo': contrato_c,
            'status': status_raw,
            'data_outorga': None,
            'qtd_outorgada': 0,
        }
    c = contratos_raw[key]
    if operacao and 'outorga' in operacao.lower():
        c['data_outorga'] = data_op
        c['qtd_outorgada'] = abs(float(qtd)) if qtd else 0
    # status final
    if status_raw == 'Finalizado':
        c['status'] = 'cancelado'
    elif status_raw == 'Em curso':
        c['status'] = 'ativo'

existing_contratos = sb.table('contratos_equity').select('id,clicksign_envelope_id').eq('organizacao_id', ORG_ID).execute().data
existing_codigos = {c['clicksign_envelope_id'] for c in existing_contratos if c.get('clicksign_envelope_id')}

contratos_inseridos = 0
for (contrato_c, benef_nome, prog_nome), c in contratos_raw.items():
    if contrato_c in existing_codigos: continue
    benef_id = pessoas_ids.get(benef_nome)
    if not benef_id:
        print(f"  ! Beneficiário não encontrado: {benef_nome}")
        continue
    prog_id = prog_ids.get(c['prog'])
    try:
        sb.table('contratos_equity').insert({
            'organizacao_id': ORG_ID,
            'plano_id': plano_id,
            'programa_id': prog_id,
            'beneficiario_id': benef_id,
            'tipo': 'stock_options',
            'status': c['status'],
            'quantidade_outorgada': c['qtd_outorgada'],
            'preco_exercicio_strike': 0,
            'natureza': 'mercantil',
            'data_aprovacao': c['data_outorga'][:10] if c['data_outorga'] else None,
            'data_assinatura': c['data_outorga'][:10] if c['data_outorga'] else None,
            'clicksign_envelope_id': contrato_c,  # guarda o código SO como referência
        }).execute()
        contratos_inseridos += 1
    except Exception as e:
        print(f"  ! Contrato {contrato_c}: {e}")

print(f"✓ Contratos equity: {contratos_inseridos} inseridos")
print("\n✅ Importação concluída!")
